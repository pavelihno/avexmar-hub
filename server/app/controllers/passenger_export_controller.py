from io import BytesIO

from flask import request, jsonify, send_file
from openpyxl import Workbook

from app.models.flight import Flight
from app.models.booking_flight import BookingFlight
from app.models.route import Route
from app.utils.datetime import format_date
from app.middlewares.auth_middleware import admin_required

@admin_required
def get_flight_passenger_export(current_user):
    flight_id = request.args.get('flight_id', type=int)
    flight_date = request.args.get('date')
    if not flight_id or not flight_date:
        return jsonify({'message': 'flight_id and date are required'}), 400

    flight = Flight.get_or_404(flight_id)
    if format_date(flight.scheduled_departure) != format_date(flight_date):
        return jsonify({'message': 'Flight date mismatch'}), 400

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Рейс:'
    ws['B1'] = flight.airline_flight_number
    ws['D1'] = 'Дата рейса:'
    ws['E1'] = format_date(flight.scheduled_departure)
    route = flight.route
    origin = route.origin_airport.city_name
    dest = route.destination_airport.city_name
    ws['G1'] = 'Маршрут:'
    ws['H1'] = f'{origin} - {dest}'

    ws.append([])
    ws.append(
        [
            '№',
            'Фамилия, Имя',
            'Пол',
            'Дата рождения',
            '№ паспорта',
            'Срок действия паспорта',
            'Гражданство',
            'Количество мест/Ребенок пассажир',
            'GSP',
            'Группа',
            'Телефон',
            'E-mail',
        ]
    )

    row = 4
    booking_flights = BookingFlight.query.filter_by(flight_id=flight_id).all()
    for bf in booking_flights:
        booking = bf.booking
        phone = booking.phone_number
        email = booking.email_address
        for bp in booking.booking_passengers:
            p = bp.passenger
            ws.cell(row=row, column=1, value=row - 3)
            ws.cell(row=row, column=2, value=f'{p.last_name} {p.first_name}')
            ws.cell(row=row, column=3, value=p.gender.value)
            ws.cell(row=row, column=4, value=format_date(p.birth_date))
            ws.cell(row=row, column=5, value=p.document_number)
            ws.cell(
                row=row,
                column=6,
                value=format_date(p.document_expiry_date),
            )
            ws.cell(row=row, column=7, value=p.citizenship.code_a2)
            ws.cell(row=row, column=11, value=phone)
            ws.cell(row=row, column=12, value=email)
            row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='flight_passengers.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@admin_required
def get_passenger_export_routes(current_user):
    routes = Route.query.join(Route.flights).distinct().all()
    data = [
        {
            'id': r.id,
            'name': f'{r.origin_airport.city_name} - {r.destination_airport.city_name}',
        }
        for r in routes
    ]
    return jsonify({'data': data})


@admin_required
def get_passenger_export_flights(current_user, route_id):
    flights = (
        Flight.query.filter_by(route_id=route_id)
        .order_by(Flight.scheduled_departure)
        .all()
    )
    data = [
        {
            'id': f.id,
            'airline_flight_number': f.airline_flight_number,
            'scheduled_departure': f.scheduled_departure.isoformat()
            if f.scheduled_departure
            else None,
        }
        for f in flights
    ]
    return jsonify({'data': data})
