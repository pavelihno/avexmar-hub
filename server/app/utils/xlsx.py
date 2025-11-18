import enum
import re
from io import BytesIO
from typing import Dict, List, Type, Tuple
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, Enum as SAEnum, Integer, Float, Numeric, Date, Time, String, Text

from app.constants.messages import XlsxMessages
from app.utils.datetime import WRITE_DATE_FORMAT, WRITE_TIME_FORMAT, format_date, format_time, parse_date_formats, parse_time_formats


def is_xlsx_file(file) -> bool:
    if not file or not getattr(file, 'filename', ''):
        return False
    extension = Path(file.filename).suffix.lower().lstrip('.')
    return extension == 'xlsx'


def get_xlsx_styles():
    font = Font(bold=True, size=12)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, border


def analyze_model_fields(model_class: Type, fields: Dict[str, str]) -> Dict[str, List]:
    enum_fields = []
    enum_values = {}
    date_fields = []
    time_fields = []
    numeric_fields = []
    text_fields = []

    mapper = inspect(model_class)

    for column in mapper.columns:
        if column.key not in fields.keys():
            continue

        col_type = column.type

        if isinstance(col_type, SAEnum):
            enum_class = col_type.enum_class
            if enum_class and issubclass(enum_class, enum.Enum):
                enum_fields.append(column.key)
                enum_values[column.key] = [e.value for e in enum_class]
        elif isinstance(col_type, Date):
            date_fields.append(column.key)
        elif isinstance(col_type, Time):
            time_fields.append(column.key)
        elif isinstance(col_type, (Integer, Float, Numeric)):
            numeric_fields.append(column.key)
        elif isinstance(col_type, (String, Text)):
            text_fields.append(column.key)

    return {
        'enum_fields': enum_fields,
        'enum_values': enum_values,
        'date_fields': date_fields,
        'time_fields': time_fields,
        'numeric_fields': numeric_fields,
        'text_fields': text_fields,
    }


"""
Methods specific for uploading data from XLSX files into models
"""


def _expand_external_fields(external_fields: Dict[Type, Dict[str, str]], max_count: int) -> dict:
    expanded = {}
    for i in range(1, max_count + 1):
        for model_class, fields_dict in external_fields.items():
            for key, label in fields_dict.items():
                expanded[f'{key}_{i}'] = f'{label} {i}'
    return expanded


def _extract_external_field(key: str) -> Tuple[str, int]:
    match = re.match(r'^(.+?)_(\d+)$', key)
    base_key = match.group(1) if match else key
    index = int(match.group(2)) if match else 0
    return base_key, index


def _get_upload_xlsx_template_wb(
    fields: dict,
    model_class: Type,
    required_fields: list = [],
    external_fields: List[Tuple[Dict[Type, Dict[str, str]], int]] = [],
    data: list = [],
) -> Workbook:
    wb = Workbook()

    all_fields = {**fields}
    external_fields_keys = set()
    field_analysis = analyze_model_fields(model_class, fields)

    # Analyze external fields if provided
    for (ext_fields_dict, _count) in external_fields:
        # Merge external fields' keys and verbose names
        expanded = _expand_external_fields(ext_fields_dict, _count)
        all_fields.update(expanded)
        external_fields_keys.update(expanded.keys())

        # Merge external fields types into analysis
        for ext_model_class, ext_fields in ext_fields_dict.items():
            ext_analysis = analyze_model_fields(ext_model_class, ext_fields)
            for base_key, values in ext_analysis.items():
                tgt = field_analysis.get(base_key)
                if isinstance(tgt, list):
                    for v in values:
                        tgt.append(v)
                elif isinstance(tgt, dict):
                    for k, v in values.items():
                        tgt[k] = v
                field_analysis[base_key] = tgt

    enum_fields = field_analysis['enum_fields']
    enum_values = field_analysis['enum_values']
    date_fields = field_analysis['date_fields']
    time_fields = field_analysis['time_fields']
    numeric_fields = field_analysis['numeric_fields']
    text_fields = field_analysis['text_fields']

    #
    # MAIN DATA SHEET
    #
    ws_data = wb.active
    ws_data.title = XlsxMessages.DATA_SHEET_NAME

    is_data_provided = len(data) > 0

    # Add header row
    if is_data_provided:
        all_fields.update({'error': 'ERROR'})
    headers = list(all_fields.values())
    ws_data.append(headers)

    # Style header
    font, border = get_xlsx_styles()
    for cell in ws_data[1]:
        cell.font = font
        cell.border = border

    # Fill data if provided
    if is_data_provided:
        for item in data:
            row_values = []

            # Add regular fields
            for key in fields.keys():
                val = item.get(key, '')
                if key in date_fields and val:
                    val = format_date(val)
                elif key in time_fields and val:
                    val = format_time(val)
                row_values.append(val if val is not None else '')

            # Add external fields
            external_data = item.get('external_data', [])
            for ext_idx, (ext_fields_dict, _count) in enumerate(external_fields):
                sub_items = external_data[ext_idx] if ext_idx < len(external_data) else []
                for i in range(_count):
                    sub_item = sub_items[i] if i < len(sub_items) else {}
                    for model_class, ext_fields in ext_fields_dict.items():
                        for key in ext_fields.keys():
                            val = sub_item.get(key, '')
                            if key in date_fields and val:
                                val = format_date(val)
                            elif key in time_fields and val:
                                val = format_time(val)
                            row_values.append(val if val is not None else '')

            # Add error column
            row_values.append(item.get('error', ''))

            ws_data.append(row_values)

    # Adjust column widths and set formatting
    for idx, base_key in enumerate(all_fields.keys(), start=1):
        col_letter = get_column_letter(idx)

        # Calculate max width based on header and data content
        if is_data_provided:
            max_len = max(
                (len(str(cell.value))
                 for cell in ws_data[col_letter] if cell.value is not None),
                default=len(str(all_fields[base_key]))
            )
        else:
            max_len = len(str(all_fields[base_key]))

        ws_data.column_dimensions[col_letter].width = max_len + 4

        # Apply cell formatting
        for cell in ws_data[col_letter]:
            if base_key in date_fields:
                cell.number_format = WRITE_DATE_FORMAT
            elif base_key in time_fields:
                cell.number_format = WRITE_TIME_FORMAT
            elif base_key in numeric_fields:
                cell.number_format = '0.00'
            else:
                cell.number_format = '@'

    #
    # RULES SHEET
    #
    ws_rules = wb.create_sheet(title=XlsxMessages.RULES_SHEET_NAME)
    ws_rules.append(
        [XlsxMessages.FIELD_COLUMN_HEADER, XlsxMessages.RULES_COLUMN_HEADER]
    )

    # Style rules header
    for cell in ws_rules[1]:
        cell.font = font
        cell.border = border

    ws_rules.column_dimensions['A'].width = 30
    ws_rules.column_dimensions['B'].width = 80

    # Adding rules for columns
    for key, display_name in all_fields.items():
        rules = []

        base_key, _ = (_extract_external_field(key)) if key in external_fields_keys else (key, 0)

        if base_key in enum_fields:
            rules.append(
                f"{XlsxMessages.ALLOWED_VALUES}: {', '.join(enum_values[base_key])}"
            )
        elif base_key in date_fields:
            rules.append(
                f"{XlsxMessages.DATE_FORMAT_LABEL}: {WRITE_DATE_FORMAT} ({XlsxMessages.DATE_FORMAT_EXAMPLE})"
            )
        elif base_key in time_fields:
            rules.append(
                f"{XlsxMessages.TIME_FORMAT_LABEL}: {WRITE_TIME_FORMAT} ({XlsxMessages.TIME_FORMAT_EXAMPLE})"
            )
        elif base_key in numeric_fields:
            rules.append(XlsxMessages.NUMERIC_VALUE)
        elif base_key in text_fields:
            rules.append(XlsxMessages.TEXT_VALUE)

        if len(rules) == 0:
            rules.append(XlsxMessages.TEXT_VALUE)

        if base_key in required_fields:
            rules = [XlsxMessages.REQUIRED_FIELD] + rules

        rule_text = '; '.join(rules)

        ws_rules.append([display_name, rule_text])

        row_idx = ws_rules.max_row
        for cell in ws_rules[row_idx]:
            cell.border = border
            if cell.column == 2:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row in ws_rules.iter_rows(min_row=2, max_row=ws_rules.max_row):
        ws_rules.row_dimensions[row[0].row].height = None

    return wb


def get_upload_xlsx_report(
    fields: dict,
    model_class: Type,
    required_fields: list = [],
    external_fields: List[Tuple[Dict[Type, Dict[str, str]], int]] = [],
    data: list = [],
) -> BytesIO:
    wb = _get_upload_xlsx_template_wb(
        fields, model_class, required_fields, external_fields, data
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_upload_xlsx_template(
    fields: dict,
    model_class: Type,
    required_fields: list = [],
    external_fields: List[Tuple[Dict[Type, Dict[str, str]], int]] = []
) -> BytesIO:

    wb = _get_upload_xlsx_template_wb(
        fields, model_class, required_fields, external_fields
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_upload_xlsx_template(
    file,
    fields: dict,
    model_class: Type,
    required_fields: list = [],
    external_fields: List[Tuple[Dict[Type, Dict[str, str]], int]] = [],
) -> list:
    all_fields = {**fields}
    external_fields_key_indices = {}
    field_analysis = analyze_model_fields(model_class, fields)

    # Analyze external fields if provided
    for i, (ext_fields_dict, _count) in enumerate(external_fields):
        # Merge external fields' keys and verbose names
        expanded = _expand_external_fields(ext_fields_dict, _count)
        all_fields.update(expanded)
        external_fields_key_indices.update({k: i for k in expanded.keys()})

        # Merge external fields types into analysis
        for ext_model_class, ext_fields in ext_fields_dict.items():
            ext_analysis = analyze_model_fields(ext_model_class, ext_fields)
            for key, values in ext_analysis.items():
                tgt = field_analysis.get(key)
                if isinstance(tgt, list):
                    for v in values:
                        tgt.append(v)

    date_fields = field_analysis['date_fields']
    time_fields = field_analysis['time_fields']

    # Parse first sheet
    wb = load_workbook(file, read_only=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {v: k for k, v in all_fields.items()}
    idx_map = {
        i: header_map[h]
        for i, h in enumerate(headers) if h in header_map
    }

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            break

        # Precreate item with all fields to ensure presence
        item = {f: None for f in fields}

        # Add external data items
        external_data = []
        for (ext_fields_dict, _count) in external_fields:
            sub_items = []
            for i in range(_count):
                sub_item = {
                    key: None
                    for model_class, fields_dict in ext_fields_dict.items()
                    for key, label in fields_dict.items()
                }
                sub_items.append(sub_item)
            external_data.append(sub_items)

        item['external_data'] = external_data

        try:
            for i, value in enumerate(row):
                key = idx_map.get(i)

                if not key:
                    continue

                is_external = key in external_fields_key_indices
                base_key, item_index = (_extract_external_field(key)) if is_external else (key, 0)

                if base_key in date_fields:
                    value = parse_date_formats(value)
                elif base_key in time_fields:
                    value = parse_time_formats(value)

                if is_external:
                    item['external_data'][external_fields_key_indices[key]][item_index - 1][base_key] = value
                else:
                    item[base_key] = value

            if required_fields:
                missing = [
                    fields[f]
                    for f in required_fields if not item.get(f)
                ]
                if missing:
                    item['error'] = XlsxMessages.missing_required_fields(missing)

        except Exception as e:
            item['error'] = str(e.with_traceback(None))

        results.append(item)

    return results
